import os
import pandas as pd
from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

input_path = os.path.join(
    SCRIPT_DIR,
    "output_categorized(2025).csv"
)

output_path = os.path.join(
    SCRIPT_DIR,
    "output_categorized_grouped_final.xlsx"
)

df = pd.read_csv(input_path)

df.columns = (
    df.columns
    .str.strip()
    .str.replace("\\", "", regex=False)
)

pm_col = "PM Interface Element ID"
country_col = "Country"
rate_card_col = "Rate Card (€)"

hour_cols = [
    "NET_Total (Hours)",
    "DCOM_Total (Hours)",
    "DEM_Total (Hours)",
    "NET_Rework (Hours)",
    "DCOM_Rework (Hours)",
    "DEM_Rework (Hours)"
]

numeric_cols = hour_cols + [rate_card_col]

df[numeric_cols] = (
    df[numeric_cols]
    .apply(pd.to_numeric, errors="coerce")
    .fillna(0)
)

df = df[
    ~df[pm_col]
    .fillna("")
    .astype(str)
    .str.contains(
        "Official_SW_Plan_Draft",
        case=False,
        na=False
    )
].copy()

df["_PM_PREFIX"] = (
    df[pm_col]
    .fillna("")
    .astype(str)
    .str.strip()
    .str.split("_", n=1)
    .str[0]
)

blank_mask = df["_PM_PREFIX"].eq("")

df.loc[blank_mask, "_PM_PREFIX"] = (
    "__BLANK__"
    + df.loc[blank_mask].index.astype(str)
)

group_cols = [
    "_PM_PREFIX",
    country_col
]

df[hour_cols] = (
    df.groupby(
        group_cols,
        sort=False,
        dropna=False
    )[hour_cols]
    .transform("sum")
)

result = df.drop_duplicates(
    subset=group_cols,
    keep="first"
).copy()

result["NET_FINAL"] = (
    result["NET_Total (Hours)"]
    + result["NET_Rework (Hours)"]
)

result["DCOM_DEM_FINAL"] = (
    result["DCOM_Total (Hours)"]
    + result["DEM_Total (Hours)"]
    + result["DCOM_Rework (Hours)"]
    + result["DEM_Rework (Hours)"]
)

result["NET"] = (
    result["NET_FINAL"] / 156
) * result[rate_card_col]

result["DCOM&DEM"] = (
    result["DCOM_DEM_FINAL"] / 156
) * result[rate_card_col]

result.drop(
    columns=["_PM_PREFIX"],
    inplace=True
)

result.to_excel(
    output_path,
    index=False,
    startrow=1,
    engine="openpyxl"
)

workbook = load_workbook(output_path)
worksheet = workbook.active

net_col = result.columns.get_loc("NET") + 1
dcom_col = result.columns.get_loc("DCOM&DEM") + 1

worksheet.merge_cells(
    start_row=1,
    start_column=net_col,
    end_row=1,
    end_column=dcom_col
)

worksheet.cell(
    row=1,
    column=net_col
).value = "Actual (ALM)"

workbook.save(output_path)

print("Done!")
print(f"Saved to: {output_path}")
print(f"Final rows: {len(result)}")