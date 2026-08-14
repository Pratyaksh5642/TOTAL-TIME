import pandas as pd
import os
from openpyxl.styles import Alignment

def process_duplicates(filename):
    # --- Resolve absolute paths ---
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, filename)

    print(f"📂 Reading Excel file from: {file_path}")

    # --- Load the data ---
    try:
        df = pd.read_excel(file_path)
    except FileNotFoundError:
        print(f"❌ Error: Could not find '{filename}' in the folder: {script_dir}")
        return

    # --- Data Cleaning ---
    df['PM_ID_clean'] = df['PM Interface Element ID'].astype(str).str.strip()
    df['Country_clean'] = df['Country'].astype(str).str.strip()

    # --- Identify Columns ---
    hours_cols = [col for col in df.columns if '(Hours)' in col]
    other_cols = [col for col in df.columns if col not in hours_cols and col not in ['PM Interface Element ID', 'Country', 'PM_ID_clean', 'Country_clean']]

    # --- Build the aggregation dictionary ---
    agg_dict = {col: 'sum' for col in hours_cols}
    agg_dict['PM Interface Element ID'] = 'first'
    agg_dict['Country'] = 'first'
    for col in other_cols:
        agg_dict[col] = 'first'

    # --- Group and Aggregate ---
    df_grouped = df.groupby(['PM_ID_clean', 'Country_clean'], as_index=False).agg(agg_dict)

    # Reorder columns to exactly match the original data (This drops the _clean columns)
    original_columns = [col for col in df.columns if col not in ['PM_ID_clean', 'Country_clean']]
    df_grouped = df_grouped[original_columns]
    df = df[original_columns] 

    # --- Rename Column for Aggregated Data Only ---
    df_grouped = df_grouped.rename(columns={'PM Interface Element ID': 'BM_ID'})

    # ==========================================
    # --- NEW CALCULATION LOGIC ---
    # ==========================================
    # 1. Ensure Rate Card is treated as a number
    rate_card = pd.to_numeric(df_grouped['Rate Card (€)'], errors='coerce').fillna(0)
    
    # 2. Calculate "Actual (ALM)" costs (Directly multiplying Hours * Rate)
    df_grouped['NET_Actual'] = df_grouped['NET_FINAL (Hours)'] * rate_card
    df_grouped['DCOM_DSM_Actual'] = df_grouped['DCOM_DSM_FINAL (Hours)'] * rate_card
    
    # 3. Calculate "Development Cost (ALM)" costs
    df_grouped['NET_Dev'] = df_grouped['NET_DEV (Hours)'] * rate_card
    df_grouped['DCOM_DSM_Dev'] = df_grouped['DCOM_DSM_DEV (Hours)'] * rate_card

    # 4. Calculate "Rework Cost (ALM)" costs
    df_grouped['NET_Rework_Cost'] = df_grouped['NET_Rework (Hours)'] * rate_card
    df_grouped['DCOM_DSM_Rework_Cost'] = df_grouped['DCOM_DSM_REWORK_TOTAL (Hours)'] * rate_card
    
    # 5. Round monetary values to 2 decimal places
    df_grouped['NET_Actual'] = df_grouped['NET_Actual'].round(2)
    df_grouped['DCOM_DSM_Actual'] = df_grouped['DCOM_DSM_Actual'].round(2)
    df_grouped['NET_Dev'] = df_grouped['NET_Dev'].round(2)
    df_grouped['DCOM_DSM_Dev'] = df_grouped['DCOM_DSM_Dev'].round(2)
    df_grouped['NET_Rework_Cost'] = df_grouped['NET_Rework_Cost'].round(2)
    df_grouped['DCOM_DSM_Rework_Cost'] = df_grouped['DCOM_DSM_Rework_Cost'].round(2)

    # 6. Calculate "Component Contribution (ALM)" Percentages (Row-wise)
    total_actual_row = df_grouped['NET_Actual'] + df_grouped['DCOM_DSM_Actual']
    net_contrib_pct = (df_grouped['NET_Actual'] / total_actual_row * 100).fillna(0).round(2)
    dcom_dsm_contrib_pct = (df_grouped['DCOM_DSM_Actual'] / total_actual_row * 100).fillna(0).round(2)
    
    df_grouped['NET_Contrib'] = net_contrib_pct.astype(str) + '%'
    df_grouped['DCOM_DSM_Contrib'] = dcom_dsm_contrib_pct.astype(str) + '%'

    # 7. Sort the data so identical PM IDs (BM IDs) are clustered together visually
    df_grouped = df_grouped.sort_values(by=['BM_ID', 'Country']).reset_index(drop=True)

    # 8. Calculate "Region Contribution (ALM)" Percentages (Window-wise per ID)
    # Find the grand total for each BM_ID
    total_net_per_id = df_grouped.groupby('BM_ID')['NET_Actual'].transform('sum')
    total_dcom_per_id = df_grouped.groupby('BM_ID')['DCOM_DSM_Actual'].transform('sum')
    
    # Calculate % contribution of this region against the BM_ID's grand total
    region_net_pct = (df_grouped['NET_Actual'] / total_net_per_id * 100).replace([float('inf'), -float('inf')], 0).fillna(0).round(2)
    region_dcom_pct = (df_grouped['DCOM_DSM_Actual'] / total_dcom_per_id * 100).replace([float('inf'), -float('inf')], 0).fillna(0).round(2)
    
    df_grouped['Region_NET_Contrib'] = region_net_pct.astype(str) + '%'
    df_grouped['Region_DCOM_Contrib'] = region_dcom_pct.astype(str) + '%'
    # ==========================================

    # --- Save the Results back to the SAME Excel file ---
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Original Data', index=False)
        
        # Write grouped data starting from row 1 (leaves row 0 blank for our top headers)
        df_grouped.to_excel(writer, sheet_name='Aggregated Data', index=False, startrow=1)
        
        worksheet = writer.sheets['Aggregated Data']
        total_cols = len(df_grouped.columns)
        
        # --- Format "Actual (ALM)" ---
        worksheet.merge_cells(start_row=1, start_column=total_cols - 9, end_row=1, end_column=total_cols - 8)
        cell_actual = worksheet.cell(row=1, column=total_cols - 9)
        cell_actual.value = "Actual (ALM)"
        cell_actual.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.cell(row=2, column=total_cols - 9).value = "NET"
        worksheet.cell(row=2, column=total_cols - 8).value = "DCOM&DSM"

        # --- Format "Development Cost (ALM)" ---
        worksheet.merge_cells(start_row=1, start_column=total_cols - 7, end_row=1, end_column=total_cols - 6)
        cell_dev = worksheet.cell(row=1, column=total_cols - 7)
        cell_dev.value = "Development Cost (ALM)"
        cell_dev.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.cell(row=2, column=total_cols - 7).value = "NET"
        worksheet.cell(row=2, column=total_cols - 6).value = "DCOM&DSM"

        # --- Format "Rework Cost (ALM)" ---
        worksheet.merge_cells(start_row=1, start_column=total_cols - 5, end_row=1, end_column=total_cols - 4)
        cell_rework = worksheet.cell(row=1, column=total_cols - 5)
        cell_rework.value = "Rework Cost (ALM)"
        cell_rework.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.cell(row=2, column=total_cols - 5).value = "NET"
        worksheet.cell(row=2, column=total_cols - 4).value = "DCOM&DSM"

        # --- Format "Component Contribution (ALM)" ---
        worksheet.merge_cells(start_row=1, start_column=total_cols - 3, end_row=1, end_column=total_cols - 2)
        cell_contrib = worksheet.cell(row=1, column=total_cols - 3)
        cell_contrib.value = "Component Contribution (ALM)"
        cell_contrib.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.cell(row=2, column=total_cols - 3).value = "NET"
        worksheet.cell(row=2, column=total_cols - 2).value = "DCOM&DSM"

        # --- Format "Region Contribution (ALM)" ---
        worksheet.merge_cells(start_row=1, start_column=total_cols - 1, end_row=1, end_column=total_cols)
        cell_region = worksheet.cell(row=1, column=total_cols - 1)
        cell_region.value = "Region Contribution (ALM)"
        cell_region.alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.cell(row=2, column=total_cols - 1).value = "NET"
        worksheet.cell(row=2, column=total_cols).value = "DCOM&DSM"

    print(f"\n✅ Data aggregation, sorting, and all cost calculations complete!")
    print(f"💾 Updated file saved directly into: {file_path}")

if __name__ == "__main__":
    process_duplicates('Daimler_Data.xlsx')
